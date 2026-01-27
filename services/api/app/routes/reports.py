"""
Reports Routes

Endpoints for generating and exporting reports.
"""

from datetime import datetime
from typing import Annotated, Any, Optional
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.auth import RequireAuth
from app.db.models import Course, Enrollment, EnrollmentStatus, TimelineEvent, User
from app.db.session import get_db
from app.errors import RBACError
from app.jobs.tasks import report_generate
from app.rbac import can
from fastapi import APIRouter, Depends, Query, Body
from fastapi.responses import Response, StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from sqlalchemy.orm import joinedload

router = APIRouter()


# ============= Request/Response Schemas =============


class GenerateReportRequest(BaseModel):
    reportType: str = Field(
        description="Type of report: course_progress, user_activity, enrollment"
    )
    format: Optional[str] = "xlsx"
    filters: Optional[dict] = None
    recipients: Optional[list[str]] = None


class ExportMatrixRequest(BaseModel):
    search: Optional[str] = None


class ExportProgressRequest(BaseModel):
    courseId: Optional[str] = None



# ============= Endpoints =============


@router.get("/overview")
async def get_overview(
    context: RequireAuth,
    db: Annotated[AsyncSession, Depends(get_db)],
    period: str = Query("month"),
) -> dict[str, Any]:
    if not await can(db, context, "reports:read"):
        raise RBACError("reports:read")
    # Active users and never logged in
    active_users_result = await db.execute(
        select(func.count())
        .select_from(User)
        .where(User.tenant_id == context.tenant_id, User.last_login_at.is_not(None))
    )
    active_users = active_users_result.scalar() or 0
    never_logged_result = await db.execute(
        select(func.count())
        .select_from(User)
        .where(User.tenant_id == context.tenant_id, User.last_login_at.is_(None))
    )
    never_logged_in = never_logged_result.scalar() or 0
    # Enrollment stats
    total_enrollments_result = await db.execute(
        select(func.count())
        .select_from(Enrollment)
        .where(Enrollment.tenant_id == context.tenant_id)
    )
    total_enrollments = total_enrollments_result.scalar() or 0
    completed_result = await db.execute(
        select(func.count())
        .select_from(Enrollment)
        .where(
            Enrollment.tenant_id == context.tenant_id,
            Enrollment.status == EnrollmentStatus.COMPLETED,
        )
    )
    completed_courses = completed_result.scalar() or 0
    # Learning structure counts
    from app.db.models import Branch, Category, Group, LearningPath

    course_count = (
        await db.execute(select(func.count()).select_from(Course))
    ).scalar() or 0
    category_count = (
        await db.execute(select(func.count()).select_from(Category))
    ).scalar() or 0
    branch_count = (
        await db.execute(select(func.count()).select_from(Branch))
    ).scalar() or 0
    group_count = (
        await db.execute(select(func.count()).select_from(Group))
    ).scalar() or 0
    lp_count = (
        await db.execute(select(func.count()).select_from(LearningPath))
    ).scalar() or 0
    # Activity placeholder arrays
    labels = ["Week 1", "Week 2", "Week 3", "Week 4"]
    logins = [active_users // 4] * 4
    completions = [completed_courses // 4] * 4
    # Enrollment distribution
    in_progress_result = await db.execute(
        select(func.count())
        .select_from(Enrollment)
        .where(Enrollment.status == EnrollmentStatus.IN_PROGRESS)
    )
    not_started_result = await db.execute(
        select(func.count())
        .select_from(Enrollment)
        .where(Enrollment.status == EnrollmentStatus.NOT_STARTED)
    )
    enrollment_distribution = {
        "completed": completed_courses,
        "inProgress": in_progress_result.scalar() or 0,
        "notStarted": not_started_result.scalar() or 0,
    }
    # User engagement placeholders
    user_engagement = {
        "dailyActiveUsers": active_users,
        "weeklyActiveUsers": active_users,
        "avgCompletionDays": 7,
        "certificatesIssued": 0,
    }
    return {
        "overview": {
            "activeUsers": active_users,
            "neverLoggedIn": never_logged_in,
            "assignedCourses": total_enrollments,
            "completedCourses": completed_courses,
        },
        "learningStructure": {
            "courses": course_count,
            "categories": category_count,
            "branches": branch_count,
            "groups": group_count,
            "learningPaths": lp_count,
        },
        "activity": {
            "labels": labels,
            "logins": logins,
            "completions": completions,
        },
        "enrollmentDistribution": enrollment_distribution,
        "userEngagement": user_engagement,
    }


@router.get("/dashboard")
async def get_dashboard_stats(
    context: RequireAuth,
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    """
    GET /api/reports/dashboard
    Get dashboard statistics.
    """
    if not await can(db, context, "reports:read"):
        raise RBACError("reports:read")

    # Get user count
    user_count_result = await db.execute(
        select(func.count())
        .select_from(User)
        .where(User.tenant_id == context.tenant_id)
    )
    total_users = user_count_result.scalar() or 0

    # Get active users (logged in within last 30 days)
    # Use timezone-naive datetime to match database TIMESTAMP WITHOUT TIME ZONE
    thirty_days_ago = datetime.now().replace(
        day=1, hour=0, minute=0, second=0, microsecond=0
    )
    active_users_result = await db.execute(
        select(func.count())
        .select_from(User)
        .where(User.last_login_at >= thirty_days_ago)
    )
    active_users = active_users_result.scalar() or 0

    # Get course count
    course_count_result = await db.execute(
        select(func.count())
        .select_from(Course)
        .where(Course.tenant_id == context.tenant_id)
    )
    total_courses = course_count_result.scalar() or 0

    # Get enrollment stats
    enrollment_stats_result = await db.execute(
        select(Enrollment.status, func.count(Enrollment.id)).group_by(Enrollment.status)
    )

    enrollment_stats = {
        "total": 0,
        "completed": 0,
        "inProgress": 0,
        "notStarted": 0,
    }

    for row in enrollment_stats_result:
        status, count = row
        enrollment_stats["total"] += count
        if status == EnrollmentStatus.COMPLETED:
            enrollment_stats["completed"] = count
        elif status == EnrollmentStatus.IN_PROGRESS:
            enrollment_stats["inProgress"] = count
        elif status == EnrollmentStatus.NOT_STARTED:
            enrollment_stats["notStarted"] = count

    # Calculate completion rate
    completion_rate: float = 0
    if enrollment_stats["total"] > 0:
        completion_rate = round(
            enrollment_stats["completed"] / enrollment_stats["total"] * 100, 1
        )

    return {
        "users": {
            "total": total_users,
            "active": active_users,
        },
        "courses": {
            "total": total_courses,
        },
        "enrollments": enrollment_stats,
        "completionRate": completion_rate,
    }


@router.get("/course-progress")
async def get_course_progress_report(
    context: RequireAuth,
    db: Annotated[AsyncSession, Depends(get_db)],
    courseId: Optional[str] = Query(None, description="Filter by course ID"),
) -> dict[str, Any]:
    """
    GET /api/reports/course-progress
    Get course progress report.
    """
    if not await can(db, context, "reports:read"):
        raise RBACError("reports:read")

    # Build query
    query = (
        select(
            Course.id,
            Course.title,
            Course.code,
            func.count(Enrollment.id).label("total_enrollments"),
            func.sum(
                func.case((Enrollment.status == EnrollmentStatus.COMPLETED, 1), else_=0)
            ).label("completed"),
            func.avg(Enrollment.progress).label("avg_progress"),
        )
        .join(Enrollment, Enrollment.course_id == Course.id, isouter=True)
        .group_by(Course.id, Course.title, Course.code)
    )

    if courseId:
        query = query.where(Course.id == courseId)

    result = await db.execute(query)
    rows = result.all()

    # Transform for response
    courses_data = []
    for row in rows:
        courses_data.append(
            {
                "courseId": row[0],
                "title": row[1],
                "code": row[2],
                "totalEnrollments": row[3] or 0,
                "completed": row[4] or 0,
                "avgProgress": float(row[5]) if row[5] else 0,
                "completionRate": round((row[4] or 0) / (row[3] or 1) * 100, 1),
            }
        )

    return {"data": courses_data}


@router.get("/training-matrix")
async def get_training_matrix(
    context: RequireAuth,
    db: Annotated[AsyncSession, Depends(get_db)],
    search: Optional[str] = Query(None),
) -> dict[str, Any]:
    if not await can(db, context, "reports:read"):
        raise RBACError("reports:read")
    # Users
    user_query = select(User).order_by(User.created_at.desc())
    if search:
        user_query = user_query.where(
            (User.first_name.ilike(f"%{search}%"))
            | (User.last_name.ilike(f"%{search}%"))
            | (User.email.ilike(f"%{search}%"))
        )
    users_result = await db.execute(user_query)
    users = users_result.scalars().all()
    # Courses
    course_result = await db.execute(select(Course.id, Course.title))
    course_rows = course_result.all()
    courses = [{"id": r[0], "title": r[1]} for r in course_rows]
    # Enrollments per user
    data_users: list[dict[str, Any]] = []
    for u in users:
        enr_result = await db.execute(
            select(Enrollment, Course.title)
            .join(Course, Course.id == Enrollment.course_id, isouter=True)
            .where(Enrollment.user_id == u.id)
        )
        entries = []
        for enr, title in enr_result.all():
            entries.append(
                {
                    "courseId": enr.course_id,
                    "courseName": title or "",
                    "progress": enr.progress or 0,
                    "status": (
                        str(enr.status)
                        if hasattr(enr.status, "value")
                        else str(enr.status)
                    ),
                }
            )
        data_users.append(
            {
                "userId": u.id,
                "userName": f"{u.first_name} {u.last_name}",
                "userEmail": u.email,
                "courses": entries,
            }
        )
    return {"users": data_users, "courses": courses}


@router.post("/generate")
async def generate_report(
    request: GenerateReportRequest,
    context: RequireAuth,
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict[str, Any]:
    """
    POST /api/reports/generate
    Queue a report generation job.
    """
    if not await can(db, context, "reports:create"):
        raise RBACError("reports:create")

    # Queue the report generation task
    task = report_generate.delay(
        report_id=request.reportType,
        tenant_id=context.tenant_id or "",
        user_id=context.user_id,
        format=request.format or "xlsx",
        filters=request.filters,
        recipients=request.recipients,
    )

    return {
        "success": True,
        "taskId": task.id,
        "message": "Report generation queued",
    }


@router.get("/user-activity")
async def get_user_activity_report(
    context: RequireAuth,
    db: Annotated[AsyncSession, Depends(get_db)],
    days: int = Query(30, ge=1, le=365, description="Number of days to look back"),
) -> dict[str, Any]:
    """
    GET /api/reports/user-activity
    Get user activity report.
    """
    if not await can(db, context, "reports:read"):
        raise RBACError("reports:read")

    from datetime import timedelta

    # Use naive datetime to match database TIMESTAMP WITHOUT TIME ZONE
    cutoff = datetime.utcnow() - timedelta(days=days)

    # Get new users in period
    new_users_result = await db.execute(
        select(func.count()).select_from(User).where(User.created_at >= cutoff)
    )
    new_users = new_users_result.scalar() or 0

    # Get active users (logged in during period)
    active_users_result = await db.execute(
        select(func.count()).select_from(User).where(User.last_login_at >= cutoff)
    )
    active_users = active_users_result.scalar() or 0

    # Get new enrollments in period
    new_enrollments_result = await db.execute(
        select(func.count())
        .select_from(Enrollment)
        .where(Enrollment.created_at >= cutoff)
    )
    new_enrollments = new_enrollments_result.scalar() or 0

    # Get completions in period
    completions_result = await db.execute(
        select(func.count())
        .select_from(Enrollment)
        .where(Enrollment.completed_at >= cutoff)
    )
    completions = completions_result.scalar() or 0

    return {
        "period": f"Last {days} days",
        "newUsers": new_users,
        "activeUsers": active_users,
        "newEnrollments": new_enrollments,
        "completions": completions,
    }


def format_relative_time(dt: datetime) -> str:
    if not dt:
        return ""
    now = datetime.utcnow()
    diff = now - dt
    seconds = diff.total_seconds()

    if seconds < 60:
        return "Just now"
    minutes = int(seconds / 60)
    if minutes < 60:
        return f"{minutes} minute{'s' if minutes != 1 else ''} ago"
    hours = int(minutes / 60)
    if hours < 24:
        return f"{hours} hour{'s' if hours != 1 else ''} ago"
    days = int(hours / 24)
    if days == 1:
        return "Yesterday"
    if days < 7:
        return f"{days} days ago"
    return dt.strftime("%b %d, %Y")


@router.get("/timeline")
async def get_timeline(
    context: RequireAuth,
    db: Annotated[AsyncSession, Depends(get_db)],
    from_date: Optional[str] = Query(None, alias="from"),
    to_date: Optional[str] = Query(None, alias="to"),
    event: Optional[str] = Query(None),
    user: Optional[str] = Query(None),
    course: Optional[str] = Query(None),
) -> dict[str, Any]:
    if not await can(db, context, "reports:read"):
        raise RBACError("reports:read")

    query = (
        select(TimelineEvent)
        .options(joinedload(TimelineEvent.user), joinedload(TimelineEvent.course))
        .where(TimelineEvent.tenant_id == context.tenant_id)
        .order_by(TimelineEvent.created_at.desc())
    )

    if from_date:
        try:
            dt = datetime.fromisoformat(from_date)
            query = query.where(TimelineEvent.created_at >= dt)
        except ValueError:
            pass

    if to_date:
        try:
            # Add one day to include the end date fully if it's just a date
            dt = datetime.fromisoformat(to_date)
            dt = dt.replace(hour=23, minute=59, second=59)
            query = query.where(TimelineEvent.created_at <= dt)
        except ValueError:
            pass

    if event:
        query = query.where(TimelineEvent.event_type == event)

    if user:
        query = query.where(TimelineEvent.user_id == user)

    if course:
        query = query.where(TimelineEvent.course_id == course)

    result = await db.execute(query.limit(100))
    events = result.scalars().all()

    formatted_events = []
    for e in events:
        description = e.event_type
        # Improve description based on event type and data
        if e.event_type == "user_signin":
            description = f"{e.user.first_name} {e.user.last_name} signed in"
        elif e.event_type == "course_created":
            course_title = e.course.title if e.course else "Unknown Course"
            description = (
                f"Course '{course_title}' created by {e.user.first_name} {e.user.last_name}"
            )
        elif e.event_type == "course_completed":
            course_title = e.course.title if e.course else "Unknown Course"
            description = (
                f"{e.user.first_name} {e.user.last_name} completed '{course_title}'"
            )
        elif e.event_type == "learning_path_created":
            description = (
                f"Learning path created by {e.user.first_name} {e.user.last_name}"
            )
        # Fallback to a generic description if needed, or use data

        formatted_events.append(
            {
                "id": e.id,
                "timestamp": e.created_at,
                "relativeTime": format_relative_time(e.created_at),
                "eventType": e.event_type,
                "description": description,
            }
        )

    return {"events": formatted_events}


@router.post("/export/training-matrix")
async def export_training_matrix(
    context: RequireAuth,
    db: Annotated[AsyncSession, Depends(get_db)],
    request: ExportMatrixRequest | None = Body(None),
) -> Any:
    if not await can(db, context, "reports:read"):
        raise RBACError("reports:read")

    # Fetch Data with Outer Joins to include users without enrollments
    query = (
        select(User, Enrollment, Course)
        .outerjoin(Enrollment, Enrollment.user_id == User.id)
        .outerjoin(Course, Course.id == Enrollment.course_id)
        .where(User.tenant_id == context.tenant_id)
        .order_by(User.last_name, User.first_name, Course.title)
    )

    if request and request.search:
        search = request.search
        query = query.where(
            (User.first_name.ilike(f"%{search}%"))
            | (User.last_name.ilike(f"%{search}%"))
            | (User.email.ilike(f"%{search}%"))
            | (Course.title.ilike(f"%{search}%"))
        )

    result = await db.execute(query)
    rows = result.all()

    workbook_out = BytesIO()
    try:
        # Create Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Training Matrix"

        headers = ["User Name", "Email", "Course Code", "Course Title", "Status", "Progress (%)", "Enrollment Date"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

        row_num = 2
        for row in rows:
            # Result is a tuple based on the select: (User, Enrollment, Course)
            # If outerjoin is used, they might be None
            user, enrollment, course = row
            
            ws.cell(row=row_num, column=1, value=f"{user.first_name or ''} {user.last_name or ''}")
            ws.cell(row=row_num, column=2, value=user.email or '')
            
            if course:
                ws.cell(row=row_num, column=3, value=course.code or '')
                ws.cell(row=row_num, column=4, value=course.title or '')
            
            if enrollment:
                status_val = enrollment.status.value if hasattr(enrollment.status, "value") else enrollment.status
                ws.cell(row=row_num, column=5, value=str(status_val))
                ws.cell(row=row_num, column=6, value=enrollment.progress or 0)
                # Ensure datetime is naive or handled correctly
                created_at = enrollment.created_at
                if created_at:
                     # If datetime is aware, convert to naive (UTC) for simple excel export
                    if created_at.tzinfo:
                        created_at = created_at.replace(tzinfo=None)
                    ws.cell(row=row_num, column=7, value=created_at.strftime("%Y-%m-%d"))
            
            row_num += 1

        wb.save(workbook_out)
        workbook_out.seek(0)
    except Exception as e:
        print(f"EXPORT ERROR: {e}")
        import traceback
        traceback.print_exc()
        raise e

    return StreamingResponse(
        workbook_out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=training_matrix.xlsx"},
    )


@router.post("/export/training-progress")
async def export_training_progress(
    context: RequireAuth,
    db: Annotated[AsyncSession, Depends(get_db)],
    request: ExportProgressRequest | None = Body(None),
) -> Any:
    print("DEBUG: export_training_progress called")
    if not await can(db, context, "reports:read"):
        print("DEBUG: RBAC check failed for reports:read")
        raise RBACError("reports:read")

    print("DEBUG: Starting training progress export...")
    try:
        # Build query similar to get_course_progress_report
        query = (
            select(
                Course.title,
                Course.code,
                func.count(Enrollment.id).label("total_enrollments"),
                func.sum(
                    func.case((Enrollment.status == EnrollmentStatus.COMPLETED, 1), else_=0)
                ).label("completed"),
                func.avg(Enrollment.progress).label("avg_progress"),
            )
            .join(Enrollment, Enrollment.course_id == Course.id, isouter=True)
            .group_by(Course.id, Course.title, Course.code)
        )

        if request and request.courseId:
            query = query.where(Course.id == request.courseId)

        print("DEBUG: Executing query...")
        result = await db.execute(query)
        rows = result.all()
        print(f"DEBUG: Query returned {len(rows)} rows.")

        workbook_out = BytesIO()
        # Create Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Course Progress"

        headers = ["Course Code", "Course Title", "Total Enrollments", "Completed", "Avg Progress (%)", "Completion Rate (%)"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

        for row_num, row in enumerate(rows, 2):
            title = row[0] or ""
            code = row[1] or ""
            total = row[2] or 0
            completed = row[3] or 0
            avg_prog = float(row[4]) if row[4] else 0.0
            
            completion_rate = round((completed / total * 100), 1) if total > 0 else 0.0

            ws.cell(row=row_num, column=1, value=code)
            ws.cell(row=row_num, column=2, value=title)
            ws.cell(row=row_num, column=3, value=total)
            ws.cell(row=row_num, column=4, value=completed)
            ws.cell(row=row_num, column=5, value=round(avg_prog, 1))
            ws.cell(row=row_num, column=6, value=completion_rate)

        wb.save(workbook_out)
        workbook_out.seek(0)
        
        return StreamingResponse(
            workbook_out,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=training_progress.xlsx"},
        )
    except Exception as e:
        print(f"EXPORT ERROR (Progress): {e}")
        import traceback
        traceback.print_exc()
        raise e
